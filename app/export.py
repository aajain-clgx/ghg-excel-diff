"""
export.py
---------
Writes the QA_Log tab into a new Excel file.
The output is a standalone .xlsx with a header block followed by a
structured data table — readable without the tool.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from diff_engine import DiffResult, SEVERITY_HIGH, SEVERITY_WARN, SEVERITY_MINOR
from annotation_store import AnnotationStore


# ---------------------------------------------------------------------------
# Colour palette (Excel hex, no #)
# ---------------------------------------------------------------------------
_RED_FILL    = PatternFill("solid", fgColor="FF6B6B")
_YELLOW_FILL = PatternFill("solid", fgColor="FFD93D")
_GREEN_FILL  = PatternFill("solid", fgColor="6BCB77")
_HEADER_FILL = PatternFill("solid", fgColor="1E1E2E")
_META_FILL   = PatternFill("solid", fgColor="2A2A3E")

_SEV_FILL = {SEVERITY_HIGH: _RED_FILL, SEVERITY_WARN: _YELLOW_FILL, SEVERITY_MINOR: _GREEN_FILL}
_SEV_TEXT = {SEVERITY_HIGH: "HIGH 🔴", SEVERITY_WARN: "WARN 🟡", SEVERITY_MINOR: "MINOR 🟢"}

_THIN = Border(
    left=Side(style="thin", color="3A3A5C"),
    right=Side(style="thin", color="3A3A5C"),
    top=Side(style="thin", color="3A3A5C"),
    bottom=Side(style="thin", color="3A3A5C"),
)

# Table column definitions: (header, width, attr_on_row_dict)
_COLUMNS = [
    ("Run Date",     16),
    ("Analyst",      16),
    ("Project",      20),
    ("V1 Label",     14),
    ("V2 Label",     14),
    ("Site ID",      20),
    ("Column",       32),
    ("V1 Value",     16),
    ("V2 Value",     16),
    ("Delta %",      10),
    ("Severity",     12),
    ("Disposition",  18),
    ("Note",         40),
    ("Structural",   12),
]


def write_qa_log(
    out_path,
    diff_result: DiffResult,
    store: AnnotationStore,
    project: str,
    v1_label: str,
    v2_label: str,
):
    """
    Write the QA_Log sheet to a new Excel file.

    Returns
    -------
    Path  — resolved output path.
    """
    import io as _io
    if not isinstance(out_path, _io.IOBase):
        out_path = Path(out_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA_Log"
    ws.sheet_view.showGridLines = False

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary = diff_result.summary_sentence()

    sidecar = store.to_dict()
    v1_filename = Path(sidecar["v1"]["path"]).name
    v2_filename = Path(sidecar["v2"]["path"]).name
    threshold = sidecar.get("materiality_threshold", 5.0)

    row = 1

    # ── Meta header block ────────────────────────────────────────────────
    meta_rows = [
        ("GHG QA Log", ""),
        ("Project:", project),
        ("Analyst:", store.analyst),
        ("Run Date:", run_date),
        ("V1:", f"{v1_label} — {v1_filename}"),
        ("V2:", f"{v2_label} — {v2_filename}"),
        ("Materiality Threshold:", f"{threshold}%"),
        ("Summary:", summary),
    ]
    for key, val in meta_rows:
        ws.cell(row, 1, key).font = Font(bold=True, color="E0E0F0", size=10)
        ws.cell(row, 1).fill = _META_FILL
        ws.cell(row, 2, val).font = Font(color="E0E0F0", size=10)
        ws.cell(row, 2).fill = _META_FILL
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=len(_COLUMNS))
        row += 1

    row += 1  # blank spacer

    # ── Table header ─────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    header_row = row
    row += 1

    # ── Data rows: value-level diffs ─────────────────────────────────────
    def _write_row(data: list, severity: str | None, is_structural: bool) -> None:
        nonlocal row
        fill = _SEV_FILL.get(severity) if severity else _YELLOW_FILL
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row, col_idx, value)
            cell.font = Font(color="1E1E2E", size=10)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=(col_idx == len(_COLUMNS) - 1))
            cell.border = _THIN
        row += 1

    for site in diff_result.site_diffs:
        for cd in site.column_diffs:
            ann = store.get_annotation(site.site_id, cd.column_v1)
            disposition = ann["disposition"] if ann else ""
            ann_note = ann["note"] if ann else ""
            pct_str = f"{cd.delta_pct:+.2f}%" if cd.delta_pct is not None else ""

            _write_row([
                run_date,
                store.analyst,
                project,
                v1_label,
                v2_label,
                site.site_id,
                cd.column_v1,
                str(cd.v1_value) if cd.v1_value not in (None, "nan", "None") else "",
                str(cd.v2_value) if cd.v2_value not in (None, "nan", "None") else "",
                pct_str,
                _SEV_TEXT.get(cd.severity, cd.severity),
                disposition,
                ann_note,
                "",  # Not structural
            ], severity=cd.severity, is_structural=False)

    # ── Data rows: structural changes ─────────────────────────────────────
    for site_id in diff_result.structural.sites_added:
        _write_row([
            run_date, store.analyst, project, v1_label, v2_label,
            site_id, "—", "—", "(site added in V2)", "",
            "STRUCTURAL", "", "", "TRUE",
        ], severity=None, is_structural=True)

    for site_id in diff_result.structural.sites_removed:
        _write_row([
            run_date, store.analyst, project, v1_label, v2_label,
            site_id, "—", "(site in V1)", "—", "",
            "STRUCTURAL", "", "", "TRUE",
        ], severity=None, is_structural=True)

    for col_name in diff_result.structural.columns_added:
        _write_row([
            run_date, store.analyst, project, v1_label, v2_label,
            "—", col_name, "—", "(column added in V2)", "",
            "STRUCTURAL", "", "", "TRUE",
        ], severity=None, is_structural=True)

    for col_name in diff_result.structural.columns_removed:
        _write_row([
            run_date, store.analyst, project, v1_label, v2_label,
            "—", col_name, "(column in V1)", "—", "",
            "STRUCTURAL", "", "", "TRUE",
        ], severity=None, is_structural=True)

    # ── Freeze panes and auto-filter ─────────────────────────────────────
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(_COLUMNS))}{header_row}"
    )

    wb.save(out_path)
    return out_path
