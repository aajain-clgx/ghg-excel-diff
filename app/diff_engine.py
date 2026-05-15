"""
diff_engine.py
--------------
Core diff logic for the GHG Excel QA tool.

Loads two Excel sheets into pandas DataFrames, merges on a site ID column,
and produces a per-site, per-column diff result with severity classification.

Severity rules:
  🔴 High Impact  — any numeric column delta > materiality_threshold %
  🟡 Possible Error — blank appeared where a value existed; or an EF-like
                      column changed while neighbouring activity column did not
  🟢 Minor         — all deltas ≤ materiality_threshold %
  (unchanged sites are not included in output)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
import openpyxl


# ---------------------------------------------------------------------------
# Heuristic patterns for identifying emission-factor columns
# (used for the isolated-EF-change 🟡 signal)
# ---------------------------------------------------------------------------
EF_COLUMN_PATTERNS = re.compile(
    r"\bef\b|\bemission\s*factor\b|\bfactor\b",
    re.IGNORECASE,
)
ACTIVITY_COLUMN_PATTERNS = re.compile(
    r"\bactivity\b|\bconsumption\b|\busage\b|\bmmbtu\b|\bmwh\b|\bkwh\b|\blitres?\b|\bgallons?\b",
    re.IGNORECASE,
)

SEVERITY_HIGH = "HIGH"       # 🔴
SEVERITY_WARN = "WARN"       # 🟡
SEVERITY_MINOR = "MINOR"     # 🟢


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ColumnDiff:
    """Change in a single column for a single site."""
    column_v1: str              # Column name as it appears in V1
    column_v2: str              # Column name as it appears in V2
    v1_value: object
    v2_value: object
    delta_abs: float | None     # Numeric absolute change (None if non-numeric)
    delta_pct: float | None     # Numeric % change relative to V1 (None if non-numeric)
    severity: str               # SEVERITY_* constant
    hint: str = ""              # Human-readable hint for the analyst


@dataclass
class SiteDiff:
    """All changes for a single site."""
    site_id: str
    column_diffs: list[ColumnDiff] = field(default_factory=list)

    @property
    def severity(self) -> str:
        """Worst severity across all column diffs for this site."""
        if any(c.severity == SEVERITY_HIGH for c in self.column_diffs):
            return SEVERITY_HIGH
        if any(c.severity == SEVERITY_WARN for c in self.column_diffs):
            return SEVERITY_WARN
        return SEVERITY_MINOR

    @property
    def max_delta_pct(self) -> float:
        """Largest absolute % delta across all numeric columns."""
        pcts = [abs(c.delta_pct) for c in self.column_diffs if c.delta_pct is not None]
        return max(pcts) if pcts else 0.0


@dataclass
class StructuralChanges:
    sites_added: list[str] = field(default_factory=list)    # In V2, not V1
    sites_removed: list[str] = field(default_factory=list)  # In V1, not V2
    columns_added: list[str] = field(default_factory=list)  # V2-only columns
    columns_removed: list[str] = field(default_factory=list)  # V1-only columns


@dataclass
class DiffResult:
    site_diffs: list[SiteDiff] = field(default_factory=list)
    structural: StructuralChanges = field(default_factory=StructuralChanges)
    total_sites_v1: int = 0
    total_sites_v2: int = 0

    @property
    def changed_site_count(self) -> int:
        return len(self.site_diffs)

    @property
    def high_impact_sites(self) -> list[SiteDiff]:
        return [s for s in self.site_diffs if s.severity == SEVERITY_HIGH]

    @property
    def warn_sites(self) -> list[SiteDiff]:
        return [s for s in self.site_diffs if s.severity == SEVERITY_WARN]

    @property
    def minor_sites(self) -> list[SiteDiff]:
        return [s for s in self.site_diffs if s.severity == SEVERITY_MINOR]

    @property
    def largest_delta_site(self) -> SiteDiff | None:
        if not self.site_diffs:
            return None
        return max(self.site_diffs, key=lambda s: s.max_delta_pct)

    def summary_sentence(self) -> str:
        n_changed = self.changed_site_count
        n_total = self.total_sites_v1
        n_warn = len(self.warn_sites)
        n_structural = (
            len(self.structural.sites_added)
            + len(self.structural.sites_removed)
            + len(self.structural.columns_added)
            + len(self.structural.columns_removed)
        )
        largest = self.largest_delta_site
        if largest and largest.max_delta_pct:
            sign = "+" if (largest.max_delta_pct >= 0) else ""
            largest_part = (
                f"Largest delta: {largest.site_id} "
                f"({sign}{largest.max_delta_pct:.1f}%). "
            )
        else:
            largest_part = ""
        return (
            f"{n_changed} of {n_total} sites changed. "
            f"{largest_part}"
            f"{n_warn} possible error(s) flagged. "
            f"{n_structural} structural change(s)."
        )


# ---------------------------------------------------------------------------
# Excel loading helpers
# ---------------------------------------------------------------------------

def list_sheets(file_path: str | Path) -> list[str]:
    """Return the sheet names in an Excel file without loading all data."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def load_sheet(file_path: str | Path, sheet_name: str) -> pd.DataFrame:
    """
    Load a single sheet from an Excel file into a DataFrame.
    All values are read as-is (no type coercion yet — that happens in the diff).

    Raises
    ------
    ValueError
        If the sheet does not exist in the file.
    PermissionError
        Re-raised if the file is password-protected (openpyxl raises this).
    """
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine="openpyxl",
            dtype=str,          # Load everything as string first; parse later
            header=0,
        )
    except Exception as exc:
        msg = str(exc).lower()
        if "encrypted" in msg or "password" in msg:
            raise PermissionError(
                f"'{Path(file_path).name}' is password-protected. "
                "Open it in Excel, remove the protection "
                "(Review → Protect Sheet/Workbook), then try again."
            ) from exc
        raise
    # Strip whitespace from all string values and column names
    df.columns = [str(c).strip() for c in df.columns]
    df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
    return df


def load_template_headers(file_path) -> list[str]:
    """Return column headers from the first sheet of a template file.

    Accepts a file path (str/Path) or a file-like object (e.g. BytesIO).
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))
               if cell.value is not None]
    wb.close()
    return headers


# ---------------------------------------------------------------------------
# Diff helpers
# ---------------------------------------------------------------------------

def _try_numeric(value: str | None) -> float | None:
    """Attempt to parse a string as a float. Returns None if not numeric."""
    if value is None or str(value).strip() in ("", "nan", "None"):
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def _classify_column_diff(
    col_v1: str,
    col_v2: str,
    v1_val: object,
    v2_val: object,
    threshold: float,
    all_cols_v1: list[str],
    site_row_v1: pd.Series,
    site_row_v2: pd.Series,
) -> ColumnDiff:
    """
    Compute delta and severity for a single (site, column) change.
    """
    n1 = _try_numeric(v1_val)
    n2 = _try_numeric(v2_val)

    # --- Blank appeared or disappeared ---
    v1_blank = (v1_val is None or str(v1_val).strip() in ("", "nan", "None"))
    v2_blank = (v2_val is None or str(v2_val).strip() in ("", "nan", "None"))

    if v1_blank and not v2_blank:
        return ColumnDiff(col_v1, col_v2, v1_val, v2_val,
                          delta_abs=None, delta_pct=None,
                          severity=SEVERITY_WARN,
                          hint="Value appeared in V2 where V1 was blank — verify intent.")
    if not v1_blank and v2_blank:
        return ColumnDiff(col_v1, col_v2, v1_val, v2_val,
                          delta_abs=None, delta_pct=None,
                          severity=SEVERITY_WARN,
                          hint="Value disappeared in V2 (blank where V1 had a value). "
                               "Possible copy-paste error or deleted data.")

    # --- Both numeric ---
    if n1 is not None and n2 is not None:
        delta_abs = n2 - n1
        if n1 == 0:
            delta_pct = None   # Can't compute % change from zero base
            severity = SEVERITY_WARN if delta_abs != 0 else SEVERITY_MINOR
            hint = ("V1 value was 0 — percentage change cannot be computed. "
                    f"Absolute change: {delta_abs:+.4g}.") if delta_abs != 0 else ""
        else:
            delta_pct = (delta_abs / abs(n1)) * 100
            if abs(delta_pct) > threshold:
                severity = SEVERITY_HIGH
                # Check for isolated EF change (🟡 upgrade candidate — but HIGH wins)
                hint = _ef_change_hint(col_v1, all_cols_v1, site_row_v1, site_row_v2)
                if not hint:
                    hint = (f"Change of {delta_pct:+.1f}% exceeds the "
                            f"{threshold}% materiality threshold.")
            else:
                # Below threshold — but still check for isolated EF pattern
                ef_hint = _ef_change_hint(col_v1, all_cols_v1, site_row_v1, site_row_v2)
                if ef_hint:
                    severity = SEVERITY_WARN
                    hint = ef_hint
                else:
                    severity = SEVERITY_MINOR
                    hint = ""
        return ColumnDiff(col_v1, col_v2, v1_val, v2_val, delta_abs, delta_pct, severity, hint)

    # --- Text change ---
    if str(v1_val).lower() != str(v2_val).lower():
        return ColumnDiff(col_v1, col_v2, v1_val, v2_val,
                          delta_abs=None, delta_pct=None,
                          severity=SEVERITY_WARN,
                          hint="Text value changed. Verify this change is intentional.")

    # No change (shouldn't reach here in normal flow, but just in case)
    return ColumnDiff(col_v1, col_v2, v1_val, v2_val, 0.0, 0.0, SEVERITY_MINOR, "")


def _ef_change_hint(
    col_name: str,
    all_cols: list[str],
    row_v1: pd.Series,
    row_v2: pd.Series,
) -> str:
    """
    Return a hint string if this column looks like an emission factor that
    changed while the corresponding activity data column did NOT change.
    Returns empty string if the pattern doesn't match.
    """
    if not EF_COLUMN_PATTERNS.search(col_name):
        return ""
    # Look for an activity column that is unchanged
    activity_changed = False
    for c in all_cols:
        if ACTIVITY_COLUMN_PATTERNS.search(c):
            v1_act = _try_numeric(row_v1.get(c))
            v2_act = _try_numeric(row_v2.get(c))
            if v1_act is not None and v2_act is not None and v1_act != v2_act:
                activity_changed = True
                break
    if not activity_changed:
        return ("Emission factor changed but corresponding activity data appears unchanged. "
                "Verify the factor source and version used.")
    return ""


# ---------------------------------------------------------------------------
# Main diff function
# ---------------------------------------------------------------------------

def run_diff(
    df_v1: pd.DataFrame,
    df_v2: pd.DataFrame,
    id_column: str,
    column_mapping: list[dict],   # [{"v1": ..., "v2": ..., "status": ...}, ...]
    materiality_threshold: float = 5.0,
) -> DiffResult:
    """
    Compare two DataFrames and return a DiffResult.

    Parameters
    ----------
    df_v1 : pd.DataFrame
        Data from Version 1 sheet.
    df_v2 : pd.DataFrame
        Data from Version 2 sheet.
    id_column : str
        The column name (in V1) that uniquely identifies each site.
        Must be present in both files (possibly under a different name if
        manually mapped — pass the V1 name here; mapping handles V2 lookup).
    column_mapping : list[dict]
        Serialized ColumnMatch list from column_matcher. Only pairs with
        status "matched" or "manual" are diffed.
    materiality_threshold : float
        % threshold above which a numeric change is classified as HIGH impact.

    Returns
    -------
    DiffResult
    """
    result = DiffResult()

    # Build a V2 column lookup from the mapping
    v1_to_v2_col: dict[str, str] = {}
    v1_only_cols: list[str] = []
    v2_only_cols: list[str] = []

    for pair in column_mapping:
        status = pair.get("status", "")
        if status in ("matched", "manual"):
            v1_to_v2_col[pair["v1"]] = pair["v2"]
        elif status == "v1_only":
            v1_only_cols.append(pair["v1"])
        elif status == "v2_only":
            v2_only_cols.append(pair["v2"])

    # Determine the V2 name for the ID column
    id_col_v2 = v1_to_v2_col.get(id_column, id_column)

    # Validate ID columns exist
    if id_column not in df_v1.columns:
        raise ValueError(
            f"ID column '{id_column}' not found in V1. "
            f"Available columns: {list(df_v1.columns)}"
        )
    if id_col_v2 not in df_v2.columns:
        raise ValueError(
            f"ID column '{id_col_v2}' not found in V2. "
            f"Available columns: {list(df_v2.columns)}"
        )

    # Index both DataFrames on the ID column
    df_v1 = df_v1.set_index(id_column)
    df_v2 = df_v2.set_index(id_col_v2)

    ids_v1 = set(df_v1.index.astype(str))
    ids_v2 = set(df_v2.index.astype(str))

    result.total_sites_v1 = len(ids_v1)
    result.total_sites_v2 = len(ids_v2)

    # Structural: sites added / removed
    result.structural.sites_added = sorted(ids_v2 - ids_v1)
    result.structural.sites_removed = sorted(ids_v1 - ids_v2)
    result.structural.columns_added = v2_only_cols
    result.structural.columns_removed = v1_only_cols

    # Columns to diff (exclude the ID column itself)
    diff_cols = [(v1c, v2c) for v1c, v2c in v1_to_v2_col.items() if v1c != id_column]
    all_v1_cols = [v1c for v1c, _ in diff_cols]

    # Per-site diff for matched sites
    matched_ids = ids_v1 & ids_v2

    for site_id in matched_ids:
        row_v1 = df_v1.loc[site_id]
        row_v2 = df_v2.loc[site_id]

        # Handle duplicate site IDs gracefully: use the first occurrence
        if isinstance(row_v1, pd.DataFrame):
            row_v1 = row_v1.iloc[0]
        if isinstance(row_v2, pd.DataFrame):
            row_v2 = row_v2.iloc[0]

        site_diff = SiteDiff(site_id=str(site_id))

        for v1_col, v2_col in diff_cols:
            v1_val = row_v1.get(v1_col)
            v2_val = row_v2.get(v2_col)

            # Skip if identical (string comparison after normalization)
            v1_str = str(v1_val).strip() if v1_val is not None else ""
            v2_str = str(v2_val).strip() if v2_val is not None else ""
            if v1_str == v2_str:
                continue

            col_diff = _classify_column_diff(
                col_v1=v1_col,
                col_v2=v2_col,
                v1_val=v1_val,
                v2_val=v2_val,
                threshold=materiality_threshold,
                all_cols_v1=all_v1_cols,
                site_row_v1=row_v1,
                site_row_v2=row_v2,
            )
            site_diff.column_diffs.append(col_diff)

        if site_diff.column_diffs:
            result.site_diffs.append(site_diff)

    # Sort: HIGH first, then WARN, then MINOR; within tier by max delta desc
    severity_order = {SEVERITY_HIGH: 0, SEVERITY_WARN: 1, SEVERITY_MINOR: 2}
    result.site_diffs.sort(
        key=lambda s: (severity_order[s.severity], -s.max_delta_pct)
    )

    return result
