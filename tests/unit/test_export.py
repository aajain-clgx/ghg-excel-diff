"""
tests/unit/test_export.py
-------------------------
Unit tests for export.py — write_qa_log() coverage:
  - BytesIO and file-path output
  - QA_Log sheet structure (meta block, column headers, data rows)
  - Structural change rows
  - Annotation disposition + note surfaced correctly
  - Severity fill colours applied
  - Freeze pane + auto-filter set
  - Empty diff produces only meta + header rows
"""

import io
import sys
from pathlib import Path
from datetime import datetime

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "app"))

from export import write_qa_log, _COLUMNS
from diff_engine import (
    DiffResult, SiteDiff, ColumnDiff, StructuralChanges,
    SEVERITY_HIGH, SEVERITY_WARN, SEVERITY_MINOR,
)
from annotation_store import AnnotationStore

DISPOSITION_EXPECTED = AnnotationStore.DISPOSITION_EXPECTED
DISPOSITION_ERROR = AnnotationStore.DISPOSITION_ERROR


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures & builders
# ─────────────────────────────────────────────────────────────────────────────

def _minimal_store(tmp_path) -> AnnotationStore:
    """Return a minimal AnnotationStore to satisfy export requirements."""
    data = {
        "run_id": "test_run",
        "created": "2024-01-01T00:00:00",
        "analyst": "Test Analyst",
        "project": "Test Project",
        "v1": {"path": "/tmp/v1.xlsx", "label": "2023"},
        "v2": {"path": "/tmp/v2.xlsx", "label": "2024"},
        "id_column": "Site ID",
        "materiality_threshold": 5.0,
        "column_mapping": [],
        "structural_changes": {},
        "annotations": {},
    }
    return AnnotationStore(tmp_path / "ann.json", data)


def _column_diff(col="EF", v1="100", v2="120", sev=SEVERITY_HIGH,
                 delta_abs=20.0, delta_pct=20.0):
    return ColumnDiff(col, col, v1, v2, delta_abs, delta_pct, sev)


def _site_diff(site_id="S1", *cdiffs):
    sd = SiteDiff(site_id=site_id)
    for cd in cdiffs:
        sd.column_diffs.append(cd)
    return sd


def _build_result(site_diffs=None, sites_added=None, sites_removed=None,
                  cols_added=None, cols_removed=None):
    dr = DiffResult()
    dr.total_sites_v1 = 2
    dr.total_sites_v2 = 2
    dr.site_diffs = site_diffs or []
    dr.structural = StructuralChanges(
        sites_added=sites_added or [],
        sites_removed=sites_removed or [],
        columns_added=cols_added or [],
        columns_removed=cols_removed or [],
    )
    return dr


def _write_to_buf(tmp_path, diff_result, store, **kwargs):
    """Write to BytesIO and return (workbook, worksheet)."""
    buf = io.BytesIO()
    write_qa_log(buf, diff_result, store,
                 project=kwargs.get("project", "Test Project"),
                 v1_label=kwargs.get("v1_label", "2023"),
                 v2_label=kwargs.get("v2_label", "2024"))
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    return wb, wb["QA_Log"]


def _all_values(ws):
    """Flatten all cell values in the worksheet into a list."""
    return [cell.value for row in ws.iter_rows() for cell in row if cell.value]


def _header_row_index(ws):
    """Find the row number containing the 14 column headers."""
    for row in ws.iter_rows():
        if row[0].value == "Run Date":
            return row[0].row
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Output destination tests
# ─────────────────────────────────────────────────────────────────────────────

class TestOutputDestination:
    def test_bytesio_output_valid_xlsx(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result()
        buf = io.BytesIO()
        write_qa_log(buf, dr, store, project="P", v1_label="v1", v2_label="v2")
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert "QA_Log" in wb.sheetnames

    def test_file_path_output_creates_file(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result()
        out = tmp_path / "output.xlsx"
        write_qa_log(out, dr, store, project="P", v1_label="v1", v2_label="v2")
        assert out.exists()

    def test_file_path_output_valid_xlsx(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result()
        out = tmp_path / "output.xlsx"
        write_qa_log(out, dr, store, project="P", v1_label="v1", v2_label="v2")
        wb = openpyxl.load_workbook(out)
        assert "QA_Log" in wb.sheetnames

    def test_string_path_output(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result()
        out = str(tmp_path / "output.xlsx")
        write_qa_log(out, dr, store, project="P", v1_label="v1", v2_label="v2")
        assert Path(out).exists()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet structure
# ─────────────────────────────────────────────────────────────────────────────

class TestSheetStructure:
    def test_qa_log_sheet_exists(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert ws is not None

    def test_exactly_14_column_headers(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        hr = _header_row_index(ws)
        assert hr is not None
        headers = [ws.cell(hr, c).value for c in range(1, len(_COLUMNS) + 2)]
        # First 14 match expected names
        expected = [col[0] for col in _COLUMNS]
        assert headers[:14] == expected

    def test_column_headers_match_spec(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        hr = _header_row_index(ws)
        for idx, (name, _) in enumerate(_COLUMNS, 1):
            assert ws.cell(hr, idx).value == name

    def test_freeze_panes_set(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        # freeze_panes is a string like "A10"
        assert ws.freeze_panes is not None
        assert str(ws.freeze_panes).startswith("A")

    def test_auto_filter_set(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert ws.auto_filter.ref is not None

    def test_gridlines_hidden(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert ws.sheet_view.showGridLines is False

    def test_only_one_sheet_in_output(self, tmp_path):
        store = _minimal_store(tmp_path)
        wb, _ = _write_to_buf(tmp_path, _build_result(), store)
        assert wb.sheetnames == ["QA_Log"]


# ─────────────────────────────────────────────────────────────────────────────
# Meta header block
# ─────────────────────────────────────────────────────────────────────────────

class TestMetaBlock:
    def test_ghg_qa_log_title_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert any(cell.value == "GHG QA Log" for row in ws.iter_rows() for cell in row)

    def test_project_label_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store, project="MyProject")
        vals = _all_values(ws)
        assert "MyProject" in vals

    def test_analyst_label_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert "Test Analyst" in _all_values(ws)

    def test_v1_label_in_meta(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store, v1_label="FY2023")
        vals = _all_values(ws)
        assert any("FY2023" in str(v) for v in vals)

    def test_v2_label_in_meta(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store, v2_label="FY2024")
        vals = _all_values(ws)
        assert any("FY2024" in str(v) for v in vals)

    def test_summary_row_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        assert any(cell.value == "Summary:" for row in ws.iter_rows() for cell in row)

    def test_threshold_in_meta(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        vals = _all_values(ws)
        assert any("5.0%" in str(v) for v in vals)

    def test_meta_block_above_column_headers(self, tmp_path):
        store = _minimal_store(tmp_path)
        _, ws = _write_to_buf(tmp_path, _build_result(), store)
        ghg_row = next(
            cell.row for row in ws.iter_rows() for cell in row
            if cell.value == "GHG QA Log"
        )
        hr = _header_row_index(ws)
        assert ghg_row < hr


# ─────────────────────────────────────────────────────────────────────────────
# Data rows — value-level diffs
# ─────────────────────────────────────────────────────────────────────────────

class TestDataRows:
    def test_site_id_in_data_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("SITE-42", _column_diff())])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "SITE-42" in _all_values(ws)

    def test_column_name_in_data_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff("MyColumn"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "MyColumn" in _all_values(ws)

    def test_v1_value_in_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff(v1="999.0"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "999.0" in _all_values(ws)

    def test_v2_value_in_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff(v2="1200.0"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "1200.0" in _all_values(ws)

    def test_delta_pct_formatted_positive(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff(delta_pct=20.0))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "+20.00%" in _all_values(ws)

    def test_delta_pct_formatted_negative(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[
            _site_diff("S1", _column_diff(v1="120", v2="100", delta_pct=-16.67,
                                          delta_abs=-20, sev=SEVERITY_HIGH))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert any("-16.67%" in str(v) for v in _all_values(ws))

    def test_delta_pct_none_shows_empty(self, tmp_path):
        store = _minimal_store(tmp_path)
        cd = ColumnDiff("EF", "EF", "", "100", None, None, SEVERITY_WARN)
        dr = _build_result(site_diffs=[_site_diff("S1", cd)])
        _, ws = _write_to_buf(tmp_path, dr, store)
        # Should not raise; delta % cell should be empty or blank
        all_vals = _all_values(ws)
        assert "+None%" not in all_vals

    def test_severity_high_text_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff(sev=SEVERITY_HIGH))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert any("HIGH" in str(v) for v in _all_values(ws))

    def test_severity_warn_text_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        cd = _column_diff(sev=SEVERITY_WARN, delta_pct=3.0, delta_abs=3.0)
        dr = _build_result(site_diffs=[_site_diff("S1", cd)])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert any("WARN" in str(v) for v in _all_values(ws))

    def test_severity_minor_text_present(self, tmp_path):
        store = _minimal_store(tmp_path)
        cd = _column_diff(sev=SEVERITY_MINOR, delta_pct=1.0, delta_abs=1.0)
        dr = _build_result(site_diffs=[_site_diff("S1", cd)])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert any("MINOR" in str(v) for v in _all_values(ws))

    def test_none_v1_value_shown_as_empty(self, tmp_path):
        store = _minimal_store(tmp_path)
        cd = ColumnDiff("EF", "EF", None, "100", None, None, SEVERITY_WARN)
        dr = _build_result(site_diffs=[_site_diff("S1", cd)])
        _, ws = _write_to_buf(tmp_path, dr, store)
        # Should not contain literal "None"
        assert "None" not in _all_values(ws)

    def test_empty_diff_has_no_data_rows(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result()
        _, ws = _write_to_buf(tmp_path, dr, store)
        hr = _header_row_index(ws)
        # No rows after header row (sheet max_row should equal header_row)
        data_vals = []
        for row in ws.iter_rows(min_row=hr + 1):
            data_vals.extend(cell.value for cell in row if cell.value)
        assert data_vals == []

    def test_multiple_sites_all_written(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[
            _site_diff("S1", _column_diff()),
            _site_diff("S2", _column_diff()),
            _site_diff("S3", _column_diff()),
        ])
        _, ws = _write_to_buf(tmp_path, dr, store)
        vals = _all_values(ws)
        assert "S1" in vals
        assert "S2" in vals
        assert "S3" in vals

    def test_project_in_every_data_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff())])
        _, ws = _write_to_buf(tmp_path, dr, store, project="MyProject")
        hr = _header_row_index(ws)
        # Project is column 3
        data_row_vals = [ws.cell(hr + 1, 3).value]
        assert data_row_vals[0] == "MyProject"


# ─────────────────────────────────────────────────────────────────────────────
# Structural change rows
# ─────────────────────────────────────────────────────────────────────────────

class TestStructuralRows:
    def test_site_added_row_contains_site_id(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(sites_added=["NEW_SITE"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "NEW_SITE" in _all_values(ws)

    def test_site_added_row_structural_true(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(sites_added=["NEW_SITE"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "TRUE" in _all_values(ws)

    def test_site_removed_row_contains_site_id(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(sites_removed=["OLD_SITE"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "OLD_SITE" in _all_values(ws)

    def test_column_added_row_contains_col_name(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(cols_added=["NewColumn"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "NewColumn" in _all_values(ws)

    def test_column_removed_row_contains_col_name(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(cols_removed=["OldColumn"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "OldColumn" in _all_values(ws)

    def test_structural_label_in_severity_column(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(sites_added=["S99"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "STRUCTURAL" in _all_values(ws)

    def test_mixed_structural_and_diff_rows(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(
            site_diffs=[_site_diff("S1", _column_diff())],
            sites_added=["S99"],
        )
        _, ws = _write_to_buf(tmp_path, dr, store)
        vals = _all_values(ws)
        assert "S1" in vals
        assert "S99" in vals


# ─────────────────────────────────────────────────────────────────────────────
# Annotation surfacing
# ─────────────────────────────────────────────────────────────────────────────

class TestAnnotationInExport:
    def test_disposition_appears_in_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        store.set_annotation("S1", "EF", DISPOSITION_EXPECTED, "Looks right")
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff("EF"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert DISPOSITION_EXPECTED in _all_values(ws)

    def test_note_appears_in_row(self, tmp_path):
        store = _minimal_store(tmp_path)
        store.set_annotation("S1", "EF", DISPOSITION_EXPECTED, "My custom note")
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff("EF"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        assert "My custom note" in _all_values(ws)

    def test_unannotated_row_has_empty_disposition(self, tmp_path):
        store = _minimal_store(tmp_path)
        # No annotation set
        dr = _build_result(site_diffs=[_site_diff("S1", _column_diff("EF"))])
        _, ws = _write_to_buf(tmp_path, dr, store)
        hr = _header_row_index(ws)
        # Disposition column is #12
        disp_val = ws.cell(hr + 1, 12).value
        assert disp_val is None or disp_val == ""

    def test_different_annotations_per_site(self, tmp_path):
        store = _minimal_store(tmp_path)
        store.set_annotation("S1", "EF", DISPOSITION_EXPECTED, "ok")
        store.set_annotation("S2", "EF", DISPOSITION_ERROR, "fixing")
        dr = _build_result(site_diffs=[
            _site_diff("S1", _column_diff("EF")),
            _site_diff("S2", _column_diff("EF")),
        ])
        _, ws = _write_to_buf(tmp_path, dr, store)
        vals = _all_values(ws)
        assert DISPOSITION_EXPECTED in vals
        assert DISPOSITION_ERROR in vals


# ─────────────────────────────────────────────────────────────────────────────
# Cell fill colours
# ─────────────────────────────────────────────────────────────────────────────

class TestCellFills:
    def _data_row_first_cell(self, tmp_path, sev, delta_pct=20.0):
        store = _minimal_store(tmp_path)
        CD = ColumnDiff("EF", "EF", "100", "120", 20.0, delta_pct, sev)
        dr = _build_result(site_diffs=[_site_diff("S1", CD)])
        _, ws = _write_to_buf(tmp_path, dr, store)
        hr = _header_row_index(ws)
        return ws.cell(hr + 1, 1)

    def test_high_row_has_red_fill(self, tmp_path):
        cell = self._data_row_first_cell(tmp_path, SEVERITY_HIGH)
        assert cell.fill.fgColor.rgb[-6:] == "FF6B6B"

    def test_warn_row_has_yellow_fill(self, tmp_path):
        cell = self._data_row_first_cell(tmp_path, SEVERITY_WARN, delta_pct=3.0)
        assert cell.fill.fgColor.rgb[-6:] == "FFD93D"

    def test_minor_row_has_green_fill(self, tmp_path):
        cell = self._data_row_first_cell(tmp_path, SEVERITY_MINOR, delta_pct=1.0)
        assert cell.fill.fgColor.rgb[-6:] == "6BCB77"

    def test_structural_row_has_yellow_fill(self, tmp_path):
        store = _minimal_store(tmp_path)
        dr = _build_result(sites_added=["S99"])
        _, ws = _write_to_buf(tmp_path, dr, store)
        hr = _header_row_index(ws)
        cell = ws.cell(hr + 1, 1)
        assert cell.fill.fgColor.rgb[-6:] == "FFD93D"
