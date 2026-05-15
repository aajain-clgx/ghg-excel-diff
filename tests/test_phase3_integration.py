"""
test_phase3_integration.py
---------------------------
End-to-end integration tests for the full pipeline:
  File upload → column matching → diff → annotation → export → QA Log verification

Run with:
  python test_phase3_integration.py

All tests use the synthetic files in tests/data/.
"""

from __future__ import annotations

import io
import sys
import json
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))

import pandas as pd
import openpyxl

from column_matcher import match_columns
from diff_engine import run_diff, SEVERITY_HIGH, SEVERITY_WARN, SEVERITY_MINOR
from annotation_store import AnnotationStore, sidecar_path
from export import write_qa_log

DATA_DIR = Path(__file__).parent / "data"
V1_PATH = DATA_DIR / "ghg_v1_2024_draft1.xlsx"
V2_PATH = DATA_DIR / "ghg_v2_2024_draft2.xlsx"

PASS = "✅ PASS"
FAIL = "❌ FAIL"
_failures: list[str] = []


def check(name: str, condition: bool, detail: str = "") -> None:
    if condition:
        print(f"  {PASS}  {name}")
    else:
        msg = f"  {FAIL}  {name}" + (f"  → {detail}" if detail else "")
        print(msg)
        _failures.append(name)


def _load_xlsx(path: Path, sheet: str = "GHG Data") -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet, dtype=str, keep_default_na=False)


# ---------------------------------------------------------------------------
# Test 1 — File loading
# ---------------------------------------------------------------------------

def test_file_loading():
    print("\n── Test 1: File Loading ─────────────────────────────────────")
    df1 = _load_xlsx(V1_PATH)
    df2 = _load_xlsx(V2_PATH)

    check("V1 loads without error", df1 is not None)
    check("V2 loads without error", df2 is not None)
    check("V1 has 11 rows", len(df1) == 11, f"got {len(df1)}")
    check("V2 has 11 rows", len(df2) == 11, f"got {len(df2)}")
    check("V1 has 9 columns", len(df1.columns) == 9, f"got {len(df1.columns)}")
    check("V2 has 9 columns", len(df2.columns) == 9, f"got {len(df2.columns)}")
    check("Facility ID present in V1", "Facility ID" in df1.columns)
    check("Facility ID present in V2", "Facility ID" in df2.columns)

    return df1, df2


# ---------------------------------------------------------------------------
# Test 2 — Column Matching
# ---------------------------------------------------------------------------

def test_column_matching(df1: pd.DataFrame, df2: pd.DataFrame):
    print("\n── Test 2: Column Matching ──────────────────────────────────")
    match = match_columns(list(df1.columns), list(df2.columns))

    check("All 9 columns auto-matched", len(match.matched_pairs) == 9,
          f"matched={len(match.matched_pairs)}")
    check("No unmatched V1 columns", len(match.v1_only) == 0,
          str([m.v1_header for m in match.v1_only]))
    check("No unmatched V2 columns", len(match.v2_only) == 0,
          str([m.v2_header for m in match.v2_only]))
    check("Site ID auto-detected as 'Facility ID'",
          match.auto_id_column == "Facility ID",
          f"got '{match.auto_id_column}'")

    return match


# ---------------------------------------------------------------------------
# Test 3 — Diff Engine
# ---------------------------------------------------------------------------

def test_diff_engine(df1: pd.DataFrame, df2: pd.DataFrame, match):
    print("\n── Test 3: Diff Engine ──────────────────────────────────────")
    mapping = match.to_serializable()
    diff = run_diff(df1, df2, "Facility ID", mapping, materiality_threshold=5.0)

    # Structural
    check("FAC-010 detected as removed",
          "FAC-010" in diff.structural.sites_removed,
          str(diff.structural.sites_removed))
    check("FAC-012 detected as added",
          "FAC-012" in diff.structural.sites_added,
          str(diff.structural.sites_added))
    check("No columns removed", len(diff.structural.columns_removed) == 0)
    check("No columns added", len(diff.structural.columns_added) == 0)

    # Changed sites present
    changed_ids = {sd.site_id for sd in diff.site_diffs}
    check("FAC-001 in diffs", "FAC-001" in changed_ids)
    check("FAC-002 in diffs", "FAC-002" in changed_ids)
    check("FAC-003 in diffs", "FAC-003" in changed_ids)
    check("FAC-004 in diffs", "FAC-004" in changed_ids)
    check("FAC-007 in diffs", "FAC-007" in changed_ids)

    # Unchanged sites NOT in diffs (or have empty diffs)
    for fac in ["FAC-005", "FAC-006", "FAC-008", "FAC-009", "FAC-011"]:
        sd = next((s for s in diff.site_diffs if s.site_id == fac), None)
        has_real_diffs = sd is not None and len(sd.column_diffs) > 0
        check(f"{fac} has no column diffs", not has_real_diffs,
              f"found {len(sd.column_diffs) if sd else 0} diffs")

    # Severity classification
    fac001 = next((s for s in diff.site_diffs if s.site_id == "FAC-001"), None)
    check("FAC-001 classified HIGH",
          fac001 and fac001.severity == SEVERITY_HIGH,
          f"got {fac001.severity if fac001 else 'None'}")

    fac007 = next((s for s in diff.site_diffs if s.site_id == "FAC-007"), None)
    check("FAC-007 classified MINOR (1.7% < 5% threshold)",
          fac007 and fac007.severity == SEVERITY_MINOR,
          f"got {fac007.severity if fac007 else 'None'}")

    # Summary sentence
    summary = diff.summary_sentence()
    check("Summary sentence non-empty", bool(summary), summary)
    print(f"         Summary: \"{summary}\"")

    return diff


# ---------------------------------------------------------------------------
# Test 4 — Annotation Store
# ---------------------------------------------------------------------------

def test_annotation_store(diff):
    print("\n── Test 4: Annotation Store ─────────────────────────────────")
    with tempfile.TemporaryDirectory() as tmp:
        sidecar = sidecar_path(tmp, "2026-05-14", "Draft 1", "Draft 2")
        store = AnnotationStore.load_or_create(
            sidecar,
            analyst="QA Tester",
            project="GHG Inventory 2024",
            v1_path=str(V1_PATH),
            v1_label="Draft 1",
            v2_path=str(V2_PATH),
            v2_label="Draft 2",
            id_column="Facility ID",
            materiality_threshold=5.0,
            column_mapping={},
            structural_changes={},
        )

        # Save annotations
        store.set_annotation("FAC-001", "__all__",
                             AnnotationStore.DISPOSITION_EXPECTED,
                             "EF updated to EPA 2024 factors — confirmed correct")
        store.set_annotation("FAC-002", "__all__",
                             AnnotationStore.DISPOSITION_ERROR,
                             "Activity data error — meter re-read in progress")
        store.set_annotation("FAC-003", "__all__",
                             AnnotationStore.DISPOSITION_EXPECTED,
                             "Cascades from FAC-001 EF update — expected")

        # Verify retrieval
        ann001 = store.get_annotation("FAC-001", "__all__")
        check("FAC-001 annotation retrieved",
              ann001 is not None and ann001.get("disposition") == AnnotationStore.DISPOSITION_EXPECTED,
              str(ann001))
        check("FAC-002 annotation retrieved as ERROR",
              store.get_annotation("FAC-002", "__all__") and
              store.get_annotation("FAC-002", "__all__").get("disposition") == AnnotationStore.DISPOSITION_ERROR)
        check("FAC-001 note persisted",
              "EPA 2024" in (ann001.get("note", "") if ann001 else ""))

        # Sidecar JSON file exists and is valid
        check("Sidecar JSON written to disk", Path(sidecar).exists())
        raw = json.loads(Path(sidecar).read_text())
        check("Sidecar has analyst field", raw.get("analyst") == "QA Tester")
        check("Sidecar has 3 annotations", len(raw.get("annotations", {})) == 3,
              f"got {len(raw.get('annotations', {}))}")

        # Export gate — FAC-004 and FAC-007 unanswered HIGH
        high_keys = [
            f"{sd.site_id}::__all__"
            for sd in diff.site_diffs
            if sd.severity == SEVERITY_HIGH
        ]
        unanswered = store.unanswered_high_impact_keys(high_keys)
        check("Unanswered HIGH sites detected",
              len(unanswered) > 0,
              f"unanswered={unanswered}")
        check("FAC-001 (answered) not in unanswered",
              "FAC-001::__all__" not in unanswered)

        # Reload from disk — simulates tool restart
        store2 = AnnotationStore.load_or_create(
            sidecar,
            analyst="QA Tester",
            project="GHG Inventory 2024",
            v1_path=str(V1_PATH),
            v1_label="Draft 1",
            v2_path=str(V2_PATH),
            v2_label="Draft 2",
            id_column="Facility ID",
            materiality_threshold=5.0,
            column_mapping={},
            structural_changes={},
        )
        ann001_reload = store2.get_annotation("FAC-001", "__all__")
        check("Annotation survives store reload",
              ann001_reload and ann001_reload.get("disposition") == AnnotationStore.DISPOSITION_EXPECTED)

        return store


# ---------------------------------------------------------------------------
# Test 5 — Export
# ---------------------------------------------------------------------------

def test_export(diff, store):
    print("\n── Test 5: Export (QA Log) ──────────────────────────────────")
    buf = io.BytesIO()
    write_qa_log(buf, diff, store, "GHG Inventory 2024", "Draft 1", "Draft 2")
    buf.seek(0)

    wb = openpyxl.load_workbook(buf)
    check("Workbook has QA_Log sheet", "QA_Log" in wb.sheetnames)

    ws = wb["QA_Log"]
    all_rows = list(ws.iter_rows(values_only=True))
    check("QA_Log has rows", len(all_rows) > 5, f"got {len(all_rows)} rows")

    # Find data rows (non-empty first cell, past the header block)
    data_rows = [
        r for r in all_rows
        if r[0] and str(r[0]) not in ("GHG QA Log", "Project:", "Analyst:",
                                       "Run Date:", "V1:", "V2:",
                                       "Materiality Threshold:", "Summary:", "")
        and not str(r[0]).startswith("Facility")  # header row
    ]
    check("QA_Log has data rows", len(data_rows) > 0, f"got {len(data_rows)}")
    print(f"         {len(data_rows)} data rows in QA_Log")

    # Verify FAC-001 annotation note appears in export
    all_text = " ".join(str(cell) for row in all_rows for cell in row if cell)
    check("FAC-001 annotation note in export", "EPA 2024" in all_text)
    check("Summary sentence in export", "sites changed" in all_text.lower() or "site" in all_text.lower())

    # BytesIO download path works
    buf2 = io.BytesIO()
    write_qa_log(buf2, diff, store, "Test", "V1", "V2")
    check("BytesIO export produces non-empty output", buf2.tell() > 0)

    return True


# ---------------------------------------------------------------------------
# Test 6 — Edge Cases
# ---------------------------------------------------------------------------

def test_edge_cases():
    print("\n── Test 6: Edge Cases ───────────────────────────────────────")

    # All same data → no diffs
    df = pd.DataFrame([
        ["FAC-A", "Plant A", "East", "NG", "10000", "53.06", "530.6", "0.0", ""]
    ], columns=["Facility ID","Facility Name","Region","Fuel Type",
                "Activity Data (MMBtu)","Emission Factor (kg CO2e/MMBtu)",
                "Scope 1 Emissions (tCO2e)","Biogenic CO2 (tCO2e)","Notes"])

    match = match_columns(list(df.columns), list(df.columns))
    diff = run_diff(df, df.copy(), "Facility ID", match.to_serializable(), 5.0)
    check("Identical files → 0 diffs", len(diff.site_diffs) == 0,
          f"got {len(diff.site_diffs)} diff sites")

    # Blank notes column — no crash
    df_blank = df.copy()
    df2_blank = df.copy()
    df2_blank.at[0, "Notes"] = ""
    diff2 = run_diff(df_blank, df2_blank, "Facility ID", match.to_serializable(), 5.0)
    check("Blank value change handled without crash", True)

    # Threshold boundary — exactly at threshold
    df3a = pd.DataFrame([["FAC-B", "Plant B", "East", "NG", "10000", "53.06", "530.6", "0.0", ""]],
                         columns=df.columns)
    df3b = df3a.copy()
    df3b.at[0, "Activity Data (MMBtu)"] = str(10_000 * 1.05)  # exactly 5%
    diff3 = run_diff(df3a, df3b, "Facility ID", match.to_serializable(), 5.0)
    if diff3.site_diffs:
        sev = diff3.site_diffs[0].severity
        check("Exactly 5.0% threshold is MINOR or HIGH (boundary defined)",
              sev in (SEVERITY_MINOR, SEVERITY_HIGH))

    # Duplicate site IDs — should not crash
    df_dup = pd.DataFrame([
        ["FAC-A", "Plant A", "East", "NG", "10000", "53.06", "530.6", "0.0", ""],
        ["FAC-A", "Plant A Dup", "East", "NG", "12000", "53.06", "636.7", "0.0", ""],
    ], columns=df.columns)
    try:
        run_diff(df_dup, df_dup.copy(), "Facility ID", match.to_serializable(), 5.0)
        check("Duplicate site IDs handled without crash", True)
    except Exception as exc:
        check("Duplicate site IDs handled without crash", False, str(exc))

    # Single-row files
    diff_single = run_diff(df, df3b, "Facility ID", match.to_serializable(), 5.0)
    check("Single-row diff runs without crash", True)


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  GHG QA Tool — Phase 3 Integration Tests")
    print("=" * 60)

    if not V1_PATH.exists() or not V2_PATH.exists():
        print("\n⚠️  Test data not found. Generating...")
        import subprocess
        subprocess.run(
            [sys.executable, str(Path(__file__).parent / "make_test_data.py")],
            check=True,
        )

    df1, df2 = test_file_loading()
    match = test_column_matching(df1, df2)
    diff = test_diff_engine(df1, df2, match)

    # Build a store with some annotations for export test
    with tempfile.TemporaryDirectory() as tmp:
        sidecar = sidecar_path(tmp, "2026-05-14", "Draft 1", "Draft 2")
        store = AnnotationStore.load_or_create(
            sidecar,
            analyst="QA Tester",
            project="GHG Inventory 2024",
            v1_path=str(V1_PATH),
            v1_label="Draft 1",
            v2_path=str(V2_PATH),
            v2_label="Draft 2",
            id_column="Facility ID",
            materiality_threshold=5.0,
            column_mapping={},
            structural_changes={},
        )
        store.set_annotation("FAC-001", "__all__",
                             AnnotationStore.DISPOSITION_EXPECTED,
                             "EF updated to EPA 2024 factors — confirmed correct")
        store.set_annotation("FAC-002", "__all__",
                             AnnotationStore.DISPOSITION_ERROR,
                             "Activity data error — meter re-read in progress")

        test_annotation_store(diff)
        test_export(diff, store)

    test_edge_cases()

    print("\n" + "=" * 60)
    if _failures:
        print(f"  ❌ {len(_failures)} test(s) FAILED:")
        for f in _failures:
            print(f"     • {f}")
        sys.exit(1)
    else:
        print(f"  ✅ ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    main()
