"""
make_test_data.py
-----------------
Generates realistic GHG v1/v2 Excel test files for Phase 3 integration testing.

Files written to: tests/data/
  ghg_v1_2024_draft1.xlsx   — baseline (10 facilities, realistic GHG data)
  ghg_v2_2024_draft2.xlsx   — revised  (EF updates, activity changes, 1 new site, 1 removed)

Run with:
  python tests/make_test_data.py
"""

from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Shared column headers ─────────────────────────────────────────────────
HEADERS = [
    "Facility ID",
    "Facility Name",
    "Region",
    "Fuel Type",
    "Activity Data (MMBtu)",
    "Emission Factor (kg CO2e/MMBtu)",
    "Scope 1 Emissions (tCO2e)",
    "Biogenic CO2 (tCO2e)",
    "Notes",
]

# ── V1 data (Draft 1 — 11 facilities) ────────────────────────────────────
V1_DATA = [
    # ID         Name                     Region   Fuel    Activity   EF      Scope1   Bio     Notes
    ("FAC-001", "Riverbend Plant",        "East",  "NG",   45_200,    53.06,  2399.5,  0.0,    "Annual ops"),
    ("FAC-002", "Eastport Terminal",      "East",  "HFO",  28_400,    77.36,  2196.8,  0.0,    ""),
    ("FAC-003", "Lakewood Facility",      "West",  "NG",   61_000,    53.06,  3236.7,  0.0,    "Expansion Q3"),
    ("FAC-004", "Northgate Station",      "North", "Coal", 12_800,   103.69,  1327.2,  15.2,   "Partial year"),
    ("FAC-005", "Southfield Ops",         "South", "NG",   39_500,    53.06,  2095.9,  0.0,    ""),
    ("FAC-006", "Cedar Creek Campus",     "West",  "NG",   22_100,    53.06,  1172.6,  0.0,    ""),
    ("FAC-007", "Harbor Logistics",       "East",  "Diesel",8_900,    74.92,   666.8,  0.0,    "Fleet ops"),
    ("FAC-008", "Inland Distribution",    "North", "NG",   55_600,    53.06,  2950.1,  0.0,    ""),
    ("FAC-009", "Sunset Processing",      "West",  "Prop", 14_200,    61.46,   872.7,  0.0,    "Seasonal"),
    ("FAC-010", "Old Peaks Site",         "North", "NG",    3_100,    53.06,   164.5,  0.0,    "Decommission"),
    ("FAC-011", "Westbrook Storage",      "West",  "NG",   18_700,    53.06,   992.2,  0.0,    ""),
]

# ── V2 data (Draft 2 — changes applied) ──────────────────────────────────
# Changes:
#   FAC-001: EF updated 53.06 → 56.10 (+5.7%) — EPA 2024 factor                 [HIGH]
#   FAC-002: Activity corrected 28400 → 31200 (+9.9%) — meter re-read            [HIGH]
#   FAC-003: Scope1 recalculated (EF update cascades)                            [HIGH]
#   FAC-004: Biogenic CO2 corrected 15.2 → 0.0 — fuel reclassified              [WARN]
#   FAC-005: No change
#   FAC-006: Notes updated only (non-numeric)                                    [no diff]
#   FAC-007: Activity minor correction 8900 → 9050 (+1.7%)                      [MINOR]
#   FAC-008: No change
#   FAC-009: No change
#   FAC-010: REMOVED (decommissioned)
#   FAC-011: No change
#   FAC-012: NEW — Pinecrest Annex                                               [structural]
V2_DATA = [
    ("FAC-001", "Riverbend Plant",        "East",  "NG",   45_200,    56.10,  2535.7,  0.0,    "Annual ops — EF updated to EPA 2024"),
    ("FAC-002", "Eastport Terminal",      "East",  "HFO",  31_200,    77.36,  2413.6,  0.0,    "Activity corrected per meter re-read"),
    ("FAC-003", "Lakewood Facility",      "West",  "NG",   61_000,    56.10,  3422.1,  0.0,    "Expansion Q3 — EF updated"),
    ("FAC-004", "Northgate Station",      "North", "Coal", 12_800,   103.69,  1327.2,   0.0,   "Biogenic reclassified to 0"),
    ("FAC-005", "Southfield Ops",         "South", "NG",   39_500,    53.06,  2095.9,  0.0,    ""),
    ("FAC-006", "Cedar Creek Campus",     "West",  "NG",   22_100,    53.06,  1172.6,  0.0,    ""),
    ("FAC-007", "Harbor Logistics",       "East",  "Diesel",9_050,    74.92,   677.9,  0.0,    "Fleet ops"),
    ("FAC-008", "Inland Distribution",    "North", "NG",   55_600,    53.06,  2950.1,  0.0,    ""),
    ("FAC-009", "Sunset Processing",      "West",  "Prop", 14_200,    61.46,   872.7,  0.0,    "Seasonal"),
    ("FAC-011", "Westbrook Storage",      "West",  "NG",   18_700,    53.06,   992.2,  0.0,    ""),
    ("FAC-012", "Pinecrest Annex",        "South", "NG",    9_800,    56.10,   549.8,  0.0,    "New — acquired Q2"),
]


def _header_fill():
    return PatternFill("solid", fgColor="1F4E79")

def _header_font():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_sheet(wb: openpyxl.Workbook, data: list, sheet_name: str = "GHG Data") -> None:
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()

    ws.row_dimensions[1].height = 30

    # Write data
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = _thin()
            if col_idx in (5, 6, 7, 8):  # numeric columns
                cell.number_format = "#,##0.00"

    # Column widths
    widths = [12, 24, 8, 10, 20, 28, 24, 18, 30]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def main():
    wb1 = openpyxl.Workbook()
    _write_sheet(wb1, V1_DATA, "GHG Data")
    p1 = OUT_DIR / "ghg_v1_2024_draft1.xlsx"
    wb1.save(p1)
    print(f"✅ Written: {p1}  ({len(V1_DATA)} facilities)")

    wb2 = openpyxl.Workbook()
    _write_sheet(wb2, V2_DATA, "GHG Data")
    p2 = OUT_DIR / "ghg_v2_2024_draft2.xlsx"
    wb2.save(p2)
    print(f"✅ Written: {p2}  ({len(V2_DATA)} facilities)")

    print()
    print("Expected diff results:")
    print("  FAC-001  EF  53.06→56.10  +5.7%   🔴 HIGH")
    print("  FAC-002  Activity 28400→31200  +9.9%   🔴 HIGH")
    print("  FAC-003  Scope1 3236.7→3422.1  +5.7%   🔴 HIGH")
    print("  FAC-004  Biogenic 15.2→0.0       🟡 WARN (non-%, value→0)")
    print("  FAC-007  Activity 8900→9050  +1.7%   🟢 MINOR")
    print("  FAC-010  Removed                  🏗 structural")
    print("  FAC-012  Added                    🏗 structural")


if __name__ == "__main__":
    main()
